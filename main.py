import tkinter
from tkinter import ttk
from tkinter import messagebox
import openpyxl
import os

filepath = "C:\PythonProgramming\Tkinter_Registrador\data.xlsx"

window = tkinter.Tk() #cria a janela, widget que vai conter outros widgets
window.title("Data entry Form") #cria o titulo da janela
window.geometry("1300x750")

# frame é como se fosse uma "outra janela" dentro da janela, quadrado dentro do quadrado
Frame = tkinter.Frame(window)
#o frame está dentro da janela
Frame.pack() #quando faço o pack e o frame estiver vazio, não aparece nada pois não tem nada dentro dele

#Label frame vai existir dentro do frame

def loadData():
    if not os.path.exists(filepath): #checa se o filepath existe, se o arquivo de dados não existir ele executa
        workbook = openpyxl.Workbook() #abre um arquivo excel, prancheta
        sheet = workbook.active #pega a prancheta que está atualmente aberta
        heading = ["First name", "Last name", "Age", "Email", "Password", "Cellphone"] #titulo das colunas
        sheet.append(heading) #adiciona os titulos das colunas no excel
        workbook.save(filepath) #salva o arquivo no diretorio especificado
    
    workbook = openpyxl.load_workbook(filepath) #carrega o arquivo excel
    sheet = workbook.active #pega a página ativa do excel
    listValues = list(sheet.values) 
    #converte os valores das linhas e colunas do excel e a transforma em uma lista, por exemplo, no index 0, 
    # vai ter todos os valores da primeira linha
    print(listValues)
    
    for col_name in listValues[0]: #pega a primeira linha da tuple, lista
        dataVisualize.heading(col_name, text = col_name) #coloca no titulo da planilha
        #col_name, quando o valor da col_name, for por exemplo, First Name, eu vou adicionar o texto com o mesmo valor da col_name,
        #ou seja, vou escrever o que está no valor da col_name
    
    for value_tuple in listValues[1:]: #vai da linha 1 até o fim da lista, ou seja, da segunda linha até a ultima linha
        dataVisualize.insert('', tkinter.END, values = value_tuple)
        #começa com vazio, Cada novo item que é inserido com um valor vazio como o primeiro parâmetro 
        # será exibido como uma linha independente no widget
        #tkinter.END usa para que quando todos os valores de uma linha forem inseridos, escrever a outra
        # linha na linha de baixo, evitando a sobreescrição
        #garante que cada nova linha de dados seja inserida após todas as linhas já presentes

#def loadRecentData():
    #filepath = "C:\Python Programming\Tkinter_Registrador\data.xlsx"
    #workbook = openpyxl.load_workbook(filepath) #carrega o arquivo excel
    #sheet = workbook.active
    #listValues = list(sheet.values) #faz a listagem dos valores do arquivo excel
    #print(listValues)
    
    #for value_tuple in listValues[-1:]:
        #dataVisualize.insert('', tkinter.END, values = value_tuple)

def showPassword():
    if password_entry["show"] == '*':
        password_entry["show"] = '' #mostra vazio, ou seja, o que você está digitando
        confirm_password_entry["show"] = ''
    else:
        password_entry["show"] = '*'
        confirm_password_entry["show"] = '*'
        
def checkPassword():
    password = password_entry.get()
    confirmedPassword = confirm_password_entry.get()
    
    if password == confirmedPassword:
        return True
    
    else:
        return False

#salvando informação do usuário

def enterData():
    
    CheckTerms = checkBox_terms_var.get()
    
    password = password_entry.get()
    firstName = first_name_entry.get() #pega o conteudo que está no primeiro nome
    lastName = last_name_entry.get()
    email = email_entry.get()
    cellphone = cellphone_entry.get()
    age = age_spinbox.get()
    
    PasswordCheck = checkPassword()
    
    if CheckTerms != "Selected":
        tkinter.messagebox.showwarning(title = "Error", message = "You didnt accepted terms and conditions!")
    elif firstName == "":
        tkinter.messagebox.showerror(title="Error", message= "Please insert your First Name!")
    elif lastName == "":
        tkinter.messagebox.showerror(title="Error", message = "Please insert your Last Name")
    elif email == "":
        tkinter.messagebox.showerror(title ="Error", message = "Please insert a valid Email address.")
    elif cellphone == "":
        tkinter.messagebox.showerror(title ="Error", message = "Please insert the Cell phone number.")
    elif PasswordCheck == False:
        tkinter.messagebox.showerror(title = "Error", message = "Your passwords don't match! Please try again.")
    else:
        print ("--------------------------")
        print ("First name:", firstName)
        print ("Last name:", lastName)
        print("Age:", age)
        print ("Email:", email)
        print ("Password:", password)
        print ("Cellphone:",cellphone)
        print ("--------------------------")
        
        workbook = openpyxl.load_workbook(filepath) #abre o arquivo que está no diretório, ou seja, abre o arquivo do excel
        sheet = workbook.active
        row_values = [firstName, lastName, age, email, password, cellphone]
        sheet.append(row_values) #adiciona as variaveis na mesma linha
        workbook.save(filepath)
        
        dataVisualize.insert('', tkinter.END, values = row_values) #adiciona a nova linha no campo onde mostra 
        #a planinha excel
    
        #loadRecentData()


#layout do usuario

user_info_frame = tkinter.LabelFrame(Frame, text = "Registration", font = "Arial 9") 
#text é o titulo do labelframe
user_info_frame.grid(row = 0, column= 0, padx = 30, pady = 30)

#padx e pady especifica o tamanho do frame

#Informações do usuário
first_name_label = tkinter.Label(user_info_frame, text = "First name",font = "Arial 13")
first_name_entry = tkinter.Entry(user_info_frame, width= 30)
first_name_label.grid(row = 0, column = 0)
first_name_entry.grid(row = 1, column = 0)

last_name_label = tkinter.Label(user_info_frame, text = "Last name", font = "Arial 13")
last_name_entry = tkinter.Entry(user_info_frame, width= 30)
last_name_label.grid(row = 2, column = 0)
last_name_entry.grid(row = 3, column = 0)

email_label = tkinter.Label(user_info_frame, text = "Email", font = "Arial 13")
email_entry = tkinter.Entry(user_info_frame, width= 30 )
email_label.grid(row = 4, column = 0)
email_entry.grid(row = 5, column = 0)

password_label = tkinter.Label(user_info_frame, text = "Password", font = "Arial 13")
password_entry = tkinter.Entry(user_info_frame, width= 30, show = '*')
password_label.grid(row = 6, column = 0)
password_entry.grid(row = 7, column = 0)

confirm_password_label = tkinter.Label(user_info_frame, text = "Confirm password", font = "Arial 13")
confirm_password_entry = tkinter.Entry(user_info_frame, width= 30, show = '*')
confirm_password_label.grid(row = 8, column = 0)
confirm_password_entry.grid(row = 9, column = 0)

#Checkbox para mostrar senha
checkBox_showPassword = tkinter.Checkbutton(user_info_frame, text="show password", font="Arial 13", 
                                            command= showPassword)
checkBox_showPassword.grid(row = 10, column = 0)


#gender_label = tkinter.Label(user_info_frame, text = "Gender", font = "Arial 13")
#gender_combobox = ttk.Combobox(user_info_frame, values = ["","Male","Female","Other"]) 
#gender_label.grid(row = 10, column = 0)
#gender_combobox.grid(row = 11, column = 0)

cellphone_label = tkinter.Label(user_info_frame, text = "Cellphone", font = "Arial 13")
cellphone_entry= tkinter.Entry(user_info_frame, width= 30)
cellphone_label.grid(row = 11, column = 0)
cellphone_entry.grid(row = 12, column = 0)

#Combobox cria uma opcao com alternativas predeterminadas de resposta
#values são as opções que iremos mostrar
# Para usar o combobox temos que usar ttk

#spin box para aumentar ou diminuir numeros, é usado bastante para o usuario inserir a idade

age_label = tkinter.Label(user_info_frame, text = "Age", font = "Arial 13")
age_spinbox = tkinter.Spinbox(user_info_frame, from_ = 18, to = 110) 
#contador que vai de 18, idade minima, ate 110, idade maxima
age_label.grid(row = 13, column = 0)
age_spinbox.grid(row = 14, column = 0)

#Frame para os termos e condições 
terms_and_conditions = tkinter.LabelFrame(Frame, text = "Terms & conditions", font = "Arial 9")
terms_and_conditions.grid(row = 1, column= 0, padx = 20, pady = 5)

#guarda a informação da checkbox
checkBox_terms_var = tkinter.StringVar(value = "Not selected") #valor padrão é not selected
#cria uma variavel que guarda o valor do checkox
checkBox_Terms = tkinter.Checkbutton(terms_and_conditions, text="I acept terms and conditions", 
                                    font="Arial 11", variable= checkBox_terms_var, onvalue = "Selected", 
                                    offvalue= "Not Selected")

#onvalue = valor quando a checkbox é marcada
#offvalue = valor quando a checkbox não está marcada

checkBox_Terms.grid(row = 0, column = 0)

#Botao de registro 

#Quando o botão é pressionado executa a função
RegisterButtonFrame = tkinter.Frame(Frame)
RegisterButtonFrame.grid(row = 4, column = 0)
Register_button = tkinter.Button(RegisterButtonFrame, text = "Register", padx = 140, pady = 10, bg = "light grey", 
                                command = enterData)
Register_button.grid(row = 0, column= 0)

#Mostra a tabela excel
dataFrame = tkinter.Frame(Frame)
dataFrame.grid(row = 0, column = 1, pady = 10)
dataScrollbar = ttk.Scrollbar(dataFrame) #cria a scrollbar
dataScrollbar.pack(side = "right", fill = "y")
#a scrollbar vai ficar na direita, side = rigth
#fill = "y" vai preencher toda vertical da janela da visualização dos dados, pega todo o espaço vertical


cols = ["First name", "Last name", "Age", "Email", "Password", "Cellphone"] #lista dos titulos das colunas

dataVisualize = ttk.Treeview(dataFrame, show = "headings", columns = cols, height = 23, 
                             yscrollcommand= dataScrollbar.set) 
#yscrollcommand = dataScrollbar.set  
# Implementa a funcionalidade da scrollbar para mover o conteudo da tabela excel, de acordo com a scrollbar

#mostra o titulo das colunas, o numero de colunas é o tamanho da lista
#headings = só mostra as colunas que têm os titulos das colunas
#height mostra o quanto de linhas vai mostrar
#columns = cols neste caso vai ter 6 colunas

dataVisualize.column("First name", width= 135)
dataVisualize.column("Last name", width= 135)
dataVisualize.column("Age", width= 50)
dataVisualize.column("Email", width= 180)
dataVisualize.column("Password", width= 140)
dataVisualize.column("Cellphone", width= 140)
dataVisualize.pack()

dataScrollbar.config(command = dataVisualize.yview) 
#configurando a scrollbar para que a lista de conteúdos mova junto com a scrollbar

#espaçamento dos widgets

for widget in user_info_frame.winfo_children(): #pega todos os widgets que tem dentro do user_info_frame
    widget.grid_configure(padx = 70, pady = 5)
    
for widget in terms_and_conditions.winfo_children(): #pega todos os widgets que tem dentro do terms_and_conditions
    widget.grid_configure(padx = 55, pady = 5)

loadData()

window.mainloop() #cria um looping infinito até quando o aplicativo for fechado

